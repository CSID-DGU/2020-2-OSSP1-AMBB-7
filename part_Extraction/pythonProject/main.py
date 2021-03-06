from openpyxl import load_workbook
import openpyxl
# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("test.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']
wb = openpyxl.Workbook()
# test 엑셀의 값 계산
# Front = 정면도 = (0,x,y)
Front_0_x = load_ws['B2'].value
Front_0_y = load_ws['C2'].value
Front_1_x = load_ws['B3'].value
Front_1_y = load_ws['C3'].value
Front_2_x = load_ws['B4'].value
Front_2_y = load_ws['C4'].value
Front_3_x = load_ws['B5'].value
Front_3_y = load_ws['C5'].value
# Floor = 평면도 = (x,y,0)
Floor_0_x = load_ws['B8'].value
Floor_0_y = load_ws['C8'].value
Floor_1_x = load_ws['B9'].value
Floor_1_y = load_ws['C9'].value
Floor_2_x = load_ws['B10'].value
Floor_2_y = load_ws['C10'].value
Floor_3_x = load_ws['B11'].value
Floor_3_y = load_ws['C11'].value
# Right = 우측면도 = (x,0,y)
Right_0_x = load_ws['B14'].value
Right_0_y = load_ws['C14'].value
Right_1_x = load_ws['B15'].value
Right_1_y = load_ws['C15'].value
Right_2_x = load_ws['B16'].value
Right_2_y = load_ws['C16'].value
Right_3_x = load_ws['B17'].value
Right_3_y = load_ws['C17'].value
print("정면도")
# Front_0의 좌표 출력
print("(0", end=",")
print(Front_0_x-Front_0_x, end=",")
print(Front_0_y-Front_0_y, end=") \n")
# Front_1의 좌표 출력
print("(0", end=",")
print(Front_1_x-Front_0_x, end=",")
print(Front_1_y-Front_0_y, end=") \n")
# Front_2의 좌표 출력
print("(0", end=",")
print(Front_2_x-Front_0_x, end=",")
print(Front_2_y-Front_0_y, end=") \n")
# Front_3의 좌표 출력
print("(0", end=",")
print(Front_3_x-Front_0_x, end=",")
print(Front_3_y-Front_0_y, end=") \n")

print("평면도")
# Floor_0의 좌표 출력
print("(", end="")
print(Floor_0_x-Floor_0_x, end=",")
print(Floor_0_y-Floor_0_y, end=",")
print("0)", end="\n")
# Floor_1의 좌표 출력
print("(", end="")
print(Floor_1_x-Floor_0_x, end=",")
print(Floor_1_y-Floor_0_y, end=",")
print("0)", end="\n")
# Floor_2의 좌표 출력
print("(", end="")
print(Floor_2_x-Floor_0_x, end=",")
print(Floor_2_y-Floor_0_y, end=",")
print("0)", end="\n")
# Floor_3의 좌표 출력
print("(", end="")
print(Floor_3_x-Floor_0_x, end=",")
print(Floor_3_y-Floor_0_y, end=",")
print("0)", end="\n")

print("우측면도")
# Right_0의 좌표 출력
print("(", end="")
print(Right_0_x-Right_0_x, end=",")
print("0", end=",")
print(Right_0_y-Right_0_y, end=") \n")
# Right_1의 좌표 출력
print("(", end="")
print(Right_1_x-Right_0_x, end=",")
print("0", end=",")
print(Right_1_y-Right_0_y, end=") \n")
# Right_2의 좌표 출력
print("(", end="")
print(Right_2_x-Right_0_x, end=",")
print("0", end=",")
print(Right_2_y-Right_0_y, end=") \n")
# Right_3의 좌표 출력
print("(", end="")
print(Right_3_x-Right_0_x, end=",")
print("0", end=",")
print(Right_3_y-Right_0_y, end=") \n")

# result 엑셀 생성
wb = openpyxl.Workbook()
sheet = wb.active
sheet['B1'] = '정면도'
sheet['A2'] = 0
sheet['B2'] = Front_0_x-Front_0_x
sheet['C2'] = Front_0_y-Front_0_y
sheet['A3'] = 0
sheet['B3'] = Front_1_x-Front_0_x
sheet['C3'] = Front_1_y-Front_0_y
sheet['A4'] = 0
sheet['B4'] = Front_2_x-Front_0_x
sheet['C4'] = Front_2_y-Front_0_y
sheet['A5'] = 0
sheet['B5'] = Front_3_x-Front_0_x
sheet['C5'] = Front_3_y-Front_0_y

sheet['F1'] = '평면도'
sheet['E2'] = Floor_0_x-Floor_0_x
sheet['F2'] = Floor_0_y-Floor_0_y
sheet['G2'] = 0
sheet['E3'] = Floor_1_x-Floor_0_x
sheet['F3'] = Floor_1_y-Floor_0_y
sheet['G3'] = 0
sheet['E4'] = Floor_2_x-Floor_0_x
sheet['F4'] = Floor_2_y-Floor_0_y
sheet['G4'] = 0
sheet['E5'] = Floor_3_x-Floor_0_x
sheet['F5'] = Floor_3_y-Floor_0_y
sheet['G5'] = 0

sheet['J1'] = '우측면도'
sheet['I2'] = Right_0_x-Right_0_x
sheet['J2'] = 0
sheet['K2'] = Right_0_y-Right_0_y
sheet['I3'] = Right_1_x-Right_0_x
sheet['J3'] = 0
sheet['K3'] = Right_1_y-Right_0_y
sheet['I4'] = Right_2_x-Right_0_x
sheet['J4'] = 0
sheet['K4'] = Right_2_y-Right_0_y
sheet['I5'] = Right_3_x-Right_0_x
sheet['J5'] = 0
sheet['K5'] = Right_3_y-Right_0_y
# sheet.cell(row=3, column=3).value = '3, 3'
# sheet.append([1, 2, 3, 4, 5])
wb.save('result.xlsx')
# 셀 좌표로 값 출력
# print(load_ws.cell(1, 2).value)
